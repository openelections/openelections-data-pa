import csv
import openpyxl
import os

COUNTY = 'Erie'

OUTPUT_FILE = os.path.join('..', '2020', '20200602__pa__primary__erie__precinct.csv')
OUTPUT_HEADER = ['county', 'precinct', 'office', 'district', 'party', 'candidate', 'votes']

ERIE_FILE = os.path.join('..', '..', 'openelections-sources-pa', '2020',
                         'Erie PA 2020 General Primary Precinct Results.xlsx')

SHEET_NAME_TO_OPENELECTIONS_OFFICE_PARTY_AND_DISTRICT = {
    'President of the United States DEM': ('President', 'DEM', ''),
    'President of the United States REP': ('President', 'REP', ''),
    'Attorney General DEM': ('Attorney General', 'DEM', ''),
    'Attorney General REP': ('Attorney General', 'REP', ''),
    'Auditor General DEM': ('Auditor General', 'DEM', ''),
    'Auditor General REP': ('Auditor General', 'REP', ''),
    'State Treasurer DEM': ('State Treasurer', 'DEM', ''),
    'State Treasurer REP': ('State Treasurer', 'REP', ''),
    'Representative in Congress 16th District DEM': ('U.S. House', 'DEM', 16),
    'Representative in Congress 16th District REP': ('U.S. House', 'REP', 16),
    'Senator in the General Assembly 49th District REP': ('State Senate', 'REP', 49),
    'Senator in the General Assembly 49th District DEM': ('State Senate', 'DEM', 49),
    'Representative in the General Assembly 1st Legislative District DEM': ('General Assembly', 'DEM', 1),
    'Representative in the General Assembly 2nd Legislative District DEM': ('General Assembly', 'DEM', 2),
    'Representative in the General Assembly 3rd Legislative District DEM': ('General Assembly', 'DEM', 3),
    'Representative in the General Assembly 3rd Legislative District REP': ('General Assembly', 'REP', 3),
    'Representative in the General Assembly 4th Legislative District DEM': ('General Assembly', 'DEM', 4),
    'Representative in the General Assembly 4th Legislative District REP': ('General Assembly', 'REP', 4),
    'Representative in the General Assembly 6th Legislative District DEM': ('General Assembly', 'DEM', 6),
    'Representative in the General Assembly 6th Legislative District REP': ('General Assembly', 'REP', 6),
    'Representative in the General Assembly 17th Legislative District DEM': ('General Assembly', 'DEM', 17),
    'Representative in the General Assembly 17th Legislative District REP': ('General Assembly', 'REP', 17),
}

PARTIES = ('(REP)', '(DEM)', 'Qualified Write In')
WRITE_IN_CANDIDATE = 'Write-in'
TOTALS_PRECINCT = 'PA County - Total'

CONTEST_ROW = 2
CANDIDATE_ROW = 4
TOTAL_VOTES_PRECINCT_START_ROW = 14
MAX_ROW = 65536

CONTEST_COLUMN = 1
PRECINCT_COLUMN = 1
CANDIDATE_START_COLUMN = 3


class XLSXSheetParser:
    def __init__(self, sheet):
        self._sheet = sheet

    def _get_cell_value(self, row, column):
        return self._sheet.cell(row=row, column=column).value


class OfficeXLSXSheetParser(XLSXSheetParser):
    def __init__(self, sheet, contest):
        super().__init__(sheet)
        self._office, self._party, self._district = \
            SHEET_NAME_TO_OPENELECTIONS_OFFICE_PARTY_AND_DISTRICT[contest]
        self._candidates = list(self._get_candidates())

    def __iter__(self):
        for row_index in range(CANDIDATE_ROW + 1, MAX_ROW):
            precinct = self._get_cell_value(row_index, PRECINCT_COLUMN)
            if precinct == TOTALS_PRECINCT:
                break
            yield from self._process_precinct_row(row_index, precinct)

    def _get_candidates(self):
        column_index = CANDIDATE_START_COLUMN
        candidate = 'null'
        while candidate:
            candidate = self._get_cell_value(CANDIDATE_ROW, column_index)
            if candidate:
                candidate, *party = candidate.split('\n')
                assert not party or party[0].strip() in PARTIES
                yield candidate.strip().title()
            column_index += 1

    def _process_precinct_row(self, row_index, precinct):
        for candidate_index, candidate in enumerate(self._candidates):
            votes = self._get_cell_value(row_index, CANDIDATE_START_COLUMN + candidate_index)
            yield {'county': COUNTY, 'precinct': precinct,
                   'office': self._office, 'party': self._party, 'district': self._district,
                   'candidate': candidate, 'votes': votes}


class VotesCastXLSXSheetParser(XLSXSheetParser):
    def __iter__(self):
        for row_index in range(TOTAL_VOTES_PRECINCT_START_ROW, MAX_ROW):
            precinct = self._get_cell_value(row_index, PRECINCT_COLUMN)
            if precinct == TOTALS_PRECINCT:
                break
            votes = self._get_cell_value(row_index, PRECINCT_COLUMN + 1)
            yield {'county': COUNTY, 'precinct': precinct,
                   'office': 'Registered Voters', 'votes': votes}


def iterate_workbook(workbook):
    for sheet in workbook.worksheets:
        contest = sheet.cell(CONTEST_ROW, CONTEST_COLUMN).value
        if not contest:
            # one sheet has the contest name shifted down a row
            contest = sheet.cell(CONTEST_ROW + 1, CONTEST_COLUMN).value
        if contest:
            contest, *_ = contest.replace('\n', ' ').strip().split(' (Vote for', 1)
            if 'Delegate' not in contest:
                print(f'Processing sheet `{contest}`')
                yield from OfficeXLSXSheetParser(sheet, contest)
            else:
                print(f'Skipping sheet `{contest}`')
        else:
            print(f'Processing votes cast sheet')
            yield from VotesCastXLSXSheetParser(sheet)


def xlsx_to_csv(workbook, csv_writer):
    csv_writer.writeheader()
    for row in iterate_workbook(workbook):
        csv_writer.writerow(row)


def main():
    with open(OUTPUT_FILE, 'w', newline='') as f:
        xlsx_to_csv(openpyxl.load_workbook(ERIE_FILE),
                    csv.DictWriter(f, OUTPUT_HEADER))


if __name__ == "__main__":
    main()
