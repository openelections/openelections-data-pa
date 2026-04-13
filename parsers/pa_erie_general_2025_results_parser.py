"""Parser for Erie County PA 2025 Municipal Election SOVC spreadsheet.

The source workbook has one contest per sheet (plus Sheet1 which is a
turnout/registered-voters summary). Each contest sheet follows a consistent
layout:

  Row 1: "Page: N of 475" + timestamp
  Row 2: office name, e.g. "JUDGE OF THE SUPERIOR COURT (Vote for  1)"
  Row 4: header row with "Precinct", "Registered Voters", candidate labels,
         "Total Votes", qualified write-in labels, and "Unresolved Write-In"
  Row 5+: repeated blocks of
          <precinct name row>, "Election Day", "Mail-In", "Provisional", "Total"

Usage:
    python parsers/pa_erie_general_2025_results_parser.py \
        "<xlsx path>" 2025/counties/20251104__pa__general__erie__precinct.csv
"""

import csv
import re
import sys

import openpyxl


COUNTY = "Erie"

PARTY_MAP = {
    "DEM": "DEM",
    "REP": "REP",
    "LIBERAL": "LBR",
    "LIB": "LBR",
    "GRN": "GRN",
    "IND": "IND",
}


def _title_word(word):
    """Title-case a word, preserving ordinal suffixes ('1ST' -> '1st')."""
    m = re.match(r"^(\d+)(ST|ND|RD|TH)$", word, re.IGNORECASE)
    if m:
        return m.group(1) + m.group(2).lower()
    if word.isupper() or word.islower():
        return word.capitalize()
    return word


def title_office(raw):
    """Normalize an all-caps office string."""
    s = re.sub(r"\s+", " ", raw).strip()
    small = {"of", "the", "and", "for", "a", "an", "in", "on", "to"}
    out = []
    for i, w in enumerate(s.split(" ")):
        lw = w.lower()
        if i > 0 and lw in small:
            out.append(lw)
        else:
            out.append(_title_word(w))
    return " ".join(out)


def extract_vote_for(raw):
    m = re.search(r"\(Vote for\s+(\d+)\)", raw)
    return int(m.group(1)) if m else 1


# Regex fragments identifying parts of the office string that aren't part of
# the canonical office name.
_TERM_RE = re.compile(r"\b(?:UNEXPIRED\s+)?\d+\s*(?:-\s*)?YEAR\s+TERM\b", re.IGNORECASE)
_JUDICIAL_DISTRICT_RE = re.compile(
    r"\b\d+(?:ST|ND|RD|TH)\s+JUDICIAL\s+DISTRICT\b", re.IGNORECASE
)
_VOTE_FOR_RE = re.compile(r"\(Vote for\s+\d+\)", re.IGNORECASE)
_DISTRICT_RE = re.compile(r"\bDISTRICT\s+(\d+)\b", re.IGNORECASE)


def parse_office(raw):
    """Return (office, district, vote_for) from the raw A2 cell value."""
    vote_for = extract_vote_for(raw)
    s = _VOTE_FOR_RE.sub("", raw)

    # Split on " - " and classify each segment.
    parts = [p.strip() for p in re.split(r"\s-\s", s) if p.strip()]
    kept = []
    district = ""
    for idx, part in enumerate(parts):
        # Strip a trailing term-length phrase from within a segment.
        part_noterm = _TERM_RE.sub("", part).strip()
        part_noterm = re.sub(r"\s{2,}", " ", part_noterm)

        if idx == 0:
            kept.append(part_noterm or part)
            continue

        # Segment that names an OpenElections district: "District N".
        m = re.fullmatch(r"district\s+(\d+)", part_noterm, re.IGNORECASE)
        if m:
            district = m.group(1)
            continue

        # Segment containing a judicial district descriptor — drop entirely.
        if _JUDICIAL_DISTRICT_RE.search(part_noterm):
            continue

        # Segment that was only a term length.
        if not part_noterm:
            continue

        kept.append(part_noterm)

    cleaned = " - ".join(kept).strip()
    cleaned = _TERM_RE.sub("", cleaned).strip()
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    return title_office(cleaned), district, vote_for


def parse_candidate_header(cell):
    """Given a header cell, return (candidate_name, party, is_writein).

    Returns (None, None, None) if this is not a candidate column (Total Votes,
    Precinct, Registered Voters, None, etc.).
    """
    if cell is None:
        return None, None, None
    text = str(cell).strip()
    if not text:
        return None, None, None
    lower = text.lower()
    if lower in ("precinct", "total votes") or lower.startswith("registered"):
        return None, None, None

    # Unresolved write-in
    if "unresolved" in lower and "write" in lower:
        return "Write-ins", "", True

    # Qualified write-in: "name\nQualified Write In"
    if "qualified write in" in lower:
        name = text.split("\n")[0].strip()
        return name, "", True

    # Regular candidate: "NAME\n(PARTY)"
    m = re.match(r"(.*?)\n\s*\(([^)]+)\)", text, re.DOTALL)
    if m:
        name = re.sub(r"\s+", " ", m.group(1).strip())
        party_raw = m.group(2).strip().upper()
        party = PARTY_MAP.get(party_raw, party_raw)
        return _proper_case(name), party, False

    # Ballot-question YES/NO
    name = text.split("\n")[0].strip()
    if name.upper() in ("YES", "NO"):
        return name.capitalize(), "", False

    return _proper_case(name), "", False


def _proper_case(name):
    """Convert an ALL CAPS name to a reasonable mixed case. Leave mixed case
    names alone."""
    if not name:
        return name
    if name.isupper():
        parts = []
        for w in name.split():
            if len(w) <= 2 and w.isalpha():
                # initials / short tokens: keep as-is
                parts.append(w)
            else:
                parts.append(w.capitalize())
        return " ".join(parts)
    return name


def to_int(value):
    if value is None or value == "":
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0


def parse_turnout_sheet(ws):
    """Parse Sheet1 into precinct -> {registered, election_day, mail,
    provisional, total}."""
    out = {}
    current = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        label = row[0]
        if label is None:
            continue
        label = str(label).strip()
        reg = row[1]
        votes = row[3] if len(row) > 3 else None
        if label in ("Election Day", "Mail-In", "Provisional", "Total"):
            if current is None:
                continue
            rec = out.setdefault(current, {"registered": to_int(reg)})
            key = {
                "Election Day": "election_day",
                "Mail-In": "mail",
                "Provisional": "provisional",
                "Total": "total",
            }[label]
            rec[key] = to_int(votes)
            # also capture registered voters (same for each row)
            if reg not in (None, ""):
                rec["registered"] = to_int(reg)
        elif label.lower() in ("precinct", "county", "erie county"):
            current = None
        else:
            # Treat as a precinct header row
            current = label
    return out


def parse_contest_sheet(ws):
    """Parse a contest sheet, yielding dict rows (without county prefix)."""
    office_raw = ws.cell(2, 1).value
    if not office_raw:
        return
    office_raw = str(office_raw).strip()
    office, district, vote_for = parse_office(office_raw)

    # Build the list of candidate columns from row 4.
    max_col = ws.max_column
    header = [ws.cell(4, c).value for c in range(1, max_col + 1)]
    candidates = []  # list of (col_index_1based, name, party, is_writein)
    for idx, cell in enumerate(header, start=1):
        name, party, is_wi = parse_candidate_header(cell)
        if name is None:
            continue
        candidates.append((idx, name, party, is_wi))

    if not candidates:
        return

    # Walk rows; each precinct block is <name>, Election Day, Mail-In,
    # Provisional, Total.
    current_precinct = None
    block = {}
    for r in range(5, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None:
            continue
        label = str(label).strip()
        if label in ("County", "Erie County"):
            continue
        if label in ("Election Day", "Mail-In", "Provisional", "Total"):
            if current_precinct is None:
                continue
            block[label] = {col: ws.cell(r, col).value for col, *_ in candidates}
            if label == "Total":
                yield from _emit_block(current_precinct, office, district, vote_for, candidates, block)
                block = {}
        else:
            # New precinct header
            current_precinct = _normalize_precinct(label)
            block = {}


def _normalize_precinct(name):
    # Collapse whitespace, keep as-is otherwise
    return re.sub(r"\s+", " ", name).strip()


def _emit_block(precinct, office, district, vote_for, candidates, block):
    writein_totals = {"votes": 0, "election_day": 0, "mail": 0, "provisional": 0}
    has_writein = False
    # Stable order: regular candidates in header order, then a single Write-ins
    regular_rows = []
    for col, name, party, is_wi in candidates:
        v = to_int(block.get("Total", {}).get(col))
        ed = to_int(block.get("Election Day", {}).get(col))
        ma = to_int(block.get("Mail-In", {}).get(col))
        pr = to_int(block.get("Provisional", {}).get(col))
        if is_wi:
            writein_totals["votes"] += v
            writein_totals["election_day"] += ed
            writein_totals["mail"] += ma
            writein_totals["provisional"] += pr
            has_writein = True
        else:
            regular_rows.append({
                "county": COUNTY,
                "precinct": precinct,
                "office": office,
                "district": district,
                "party": party,
                "candidate": name,
                "votes": v,
                "election_day": ed,
                "mail": ma,
                "provisional": pr,
                "vote_for": vote_for,
            })

    yield from regular_rows
    if has_writein:
        yield {
            "county": COUNTY,
            "precinct": precinct,
            "office": office,
            "district": district,
            "party": "",
            "candidate": "Write-ins",
            "votes": writein_totals["votes"],
            "election_day": writein_totals["election_day"],
            "mail": writein_totals["mail"],
            "provisional": writein_totals["provisional"],
            "vote_for": vote_for,
        }


def main(xlsx_path, output_csv):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    turnout = parse_turnout_sheet(wb[wb.sheetnames[0]])

    fieldnames = [
        "county", "precinct", "office", "district", "party", "candidate",
        "votes", "election_day", "mail", "provisional", "vote_for",
    ]

    # Collect rows per precinct so turnout rows appear before contest rows.
    contest_rows_by_precinct = {}
    contest_order = []

    for sheet_name in wb.sheetnames[1:]:
        ws = wb[sheet_name]
        for row in parse_contest_sheet(ws):
            pct = row["precinct"]
            if pct not in contest_rows_by_precinct:
                contest_rows_by_precinct[pct] = []
                contest_order.append(pct)
            contest_rows_by_precinct[pct].append(row)

    # Preserve precinct ordering from Sheet1 if possible.
    sheet1_order = list(turnout.keys())
    ordered_precincts = [p for p in sheet1_order if p in contest_rows_by_precinct]
    for p in contest_order:
        if p not in ordered_precincts:
            ordered_precincts.append(p)

    with open(output_csv, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for pct in ordered_precincts:
            t = turnout.get(pct, {})
            if t:
                w.writerow({
                    "county": COUNTY, "precinct": pct,
                    "office": "Registered Voters", "district": "",
                    "party": "", "candidate": "",
                    "votes": t.get("registered", ""),
                    "election_day": "", "mail": "", "provisional": "",
                    "vote_for": "",
                })
                w.writerow({
                    "county": COUNTY, "precinct": pct,
                    "office": "Ballots Cast", "district": "",
                    "party": "", "candidate": "",
                    "votes": t.get("total", ""),
                    "election_day": t.get("election_day", ""),
                    "mail": t.get("mail", ""),
                    "provisional": t.get("provisional", ""),
                    "vote_for": "",
                })
            for row in contest_rows_by_precinct.get(pct, []):
                w.writerow(row)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: pa_erie_general_2025_results_parser.py <input.xlsx> <output.csv>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
