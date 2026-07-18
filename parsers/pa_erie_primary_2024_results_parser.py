#!/usr/bin/env python3
"""Parse Erie County PA 2024 Primary precinct results.

Source: Erie PA Official-Precinct-Results.xlsx
(One sheet per contest. Sheet1 is the turnout summary; Sheet2-Sheet27 are
the contests we care about. Each contest sheet has a title in row 1, column
headers in row 3, "County"/"PA County" placeholders in rows 4-5, then
precinct blocks of 5 rows: precinct name + Election Day/Mail-In/Provisional/
Total. Candidate columns are at varying positions (merged cells); we parse
row 3 to find them. Write-in columns are summed into a single Write-In Totals
entry.)

Usage:
    python parsers/pa_erie_primary_2024_results_parser.py <input.xlsx> <output.csv>
"""

import csv
import re
import sys
from pathlib import Path

import openpyxl


OFFICE_MAP = {
    "PRESIDENT OF THE UNITED STATES": ("President", False),
    "UNITED STATES SENATOR": ("U.S. Senate", False),
    "ATTORNEY GENERAL": ("Attorney General", False),
    "AUDITOR GENERAL": ("Auditor General", False),
    "STATE TREASURER": ("State Treasurer", False),
    "REPRESENTATIVE IN CONGRESS": ("U.S. House", True),
    "SENATOR IN THE GENERAL ASSEMBLY": ("State Senate", True),
    "REPRESENTATIVE IN THE GENERAL ASSEMBLY": ("State House", True),
}

DISTRICT_RE = re.compile(r"DISTRICT\s+(\d+)", re.IGNORECASE)
PARTY_RE = re.compile(r"\((DEMOCRATIC|REPUBLICAN)\)", re.IGNORECASE)

# Sheets to parse (Sheet2-Sheet27 cover the statewide + legislative contests;
# Sheet28+ are delegates/committeeperson we skip).
CONTEST_SHEET_RANGE = range(2, 28)

FIELDNAMES = [
    "county", "precinct", "office", "district", "party",
    "candidate", "votes", "election_day", "provisional", "absentee",
]


def normalize_office(raw: str) -> tuple[str, str]:
    upper = re.sub(r"\s+", " ", raw.upper()).strip()
    # strip "(VOTE FOR N)" suffix
    upper = re.sub(r"\(VOTE FOR\s+\d+\)", "", upper).strip()
    district = ""
    dm = DISTRICT_RE.search(upper)
    if dm:
        district = str(int(dm.group(1)))
        upper = DISTRICT_RE.sub("", upper).strip()
    upper = upper.rstrip("- ").strip()
    for key, (norm, extract) in OFFICE_MAP.items():
        if upper.startswith(key):
            return (norm, district if extract else "")
    return (raw.title(), district)


def finalize_candidate(name: str) -> str:
    name = name.replace(",", "").strip()
    if name.lower() == "write-in" or name.lower() == "uncommitted":
        return name.capitalize() if name.lower() == "uncommitted" else "Write-In Totals"
    parts = name.split()
    out = []
    for w in parts:
        if w.upper() in ("JR", "SR", "II", "III", "IV"):
            out.append(w.upper().replace("JR", "Jr.").replace("SR", "Sr."))
        elif "-" in w and w.upper() != w:
            out.append("-".join(p.capitalize() for p in w.split("-")))
        elif w[:2].upper() == "MC" and len(w) > 2:
            out.append("Mc" + w[2:].capitalize())
        else:
            out.append(w.capitalize())
    return " ".join(out)


def parse_candidate_header(cell: str) -> tuple[str, bool]:
    """Parse a candidate header cell. Returns (name, is_write_in).
    Cells look like 'JOSEPH R BIDEN JR\\n(DEM) ' or
    'UNCOMMITTED\\nQualified Write In \\n'."""
    if not cell:
        return ("", False)
    lines = [l.strip() for l in str(cell).split("\n") if l.strip()]
    if not lines:
        return ("", False)
    name = lines[0]
    is_write_in = len(lines) > 1 and "write in" in lines[1].lower()
    return (name, is_write_in)


def parse_sheet(ws, county: str) -> list[dict]:
    rows_out: list[dict] = []
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 6:
        return rows_out
    title_cell = all_rows[1][0]
    if not title_cell:
        return rows_out
    title = str(title_cell).split("\n")[0].strip()
    party = ""
    pm = PARTY_RE.search(title)
    if pm:
        party = "DEM" if "DEMOCRATIC" in pm.group(1).upper() else "REP"
    office_raw = re.sub(r"\s*\(DEMOCRATIC\)\s*|\s*\(REPUBLICAN\)\s*", "", title, flags=re.IGNORECASE).strip()
    office_raw = re.sub(r"\(Vote for\s+\d+\)", "", office_raw, flags=re.IGNORECASE).strip()
    office, district = normalize_office(office_raw)
    if not office or not party:
        return rows_out
    # Parse row 3 for candidate columns
    header_row = all_rows[3]
    candidate_cols: list[tuple[int, str, bool]] = []  # (col, name, is_write_in)
    for col_idx, cell in enumerate(header_row):
        if cell is None or cell == "":
            continue
        s = str(cell).strip()
        if not s:
            continue
        upper = s.upper()
        if upper in ("PRECINCT",) or upper.startswith("REGISTERED"):
            continue
        if upper == "TOTAL VOTES":
            continue
        name, is_wi = parse_candidate_header(s)
        if name:
            candidate_cols.append((col_idx, name, is_wi))
    if not candidate_cols:
        return rows_out
    # Iterate precinct blocks starting at row 6
    current_precinct = None
    for row in all_rows[6:]:
        if not row:
            continue
        label = row[0]
        if label is None:
            continue
        label_s = str(label).strip()
        if not label_s:
            continue
        if label_s in ("Election Day", "Mail-In", "Provisional", "Total"):
            if label_s != "Total" or current_precinct is None:
                continue
            # Extract votes per candidate
            for col, cand_name, is_wi in candidate_cols:
                if col >= len(row):
                    continue
                v = row[col]
                votes = 0 if v is None or v == "" else int(v)
                if is_wi:
                    out_name = "Write-In Totals"
                elif cand_name.lower() == "uncommitted":
                    out_name = "Uncommitted"
                else:
                    out_name = finalize_candidate(cand_name)
                rows_out.append({
                    "county": county, "precinct": current_precinct,
                    "office": office, "district": district,
                    "party": party, "candidate": out_name,
                    "votes": votes, "election_day": "", "provisional": "", "absentee": "",
                })
            current_precinct = None
        elif label_s in ("County", "PA County"):
            continue
        elif label_s == "Cumulative":
            current_precinct = None
            continue
        else:
            # precinct name row
            current_precinct = label_s
    # Merge write-in columns into a single Write-In Totals entry per precinct
    merged: dict[tuple[str, str], dict[str, int]] = {}
    order: list[tuple[str, str]] = []
    out: list[dict] = []
    for r in rows_out:
        key = (r["precinct"], r["candidate"])
        if r["candidate"] == "Write-In Totals":
            if key not in merged:
                merged[key] = {"votes": 0}
                order.append(key)
            merged[key]["votes"] += r["votes"]
        else:
            out.append(r)
    for key in order:
        precinct, cand = key
        out.append({
            "county": county, "precinct": precinct,
            "office": office, "district": district,
            "party": party, "candidate": cand,
            "votes": merged[key]["votes"], "election_day": "", "provisional": "", "absentee": "",
        })
    return out


def parse_erie(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    rows: list[dict] = []
    for i in CONTEST_SHEET_RANGE:
        sheet_name = f"Sheet{i}"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows.extend(parse_sheet(ws, "Erie"))
    wb.close()
    return rows


def main(argv: list[str]) -> None:
    if len(argv) != 3:
        sys.exit(f"Usage: {Path(argv[0]).name} <input.xlsx> <output.csv>")
    xlsx_path = Path(argv[1])
    out_path = Path(argv[2])
    if not xlsx_path.exists():
        sys.exit(f"Missing XLSX: {xlsx_path}")
    rows = parse_erie(xlsx_path)
    with out_path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=FIELDNAMES)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in FIELDNAMES})
    print(f"Wrote {len(rows)} rows to {out_path}")


if __name__ == "__main__":
    main(sys.argv)