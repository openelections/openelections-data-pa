#!/usr/bin/env python3
"""Parse Erie County PA 2026 Primary precinct-level results from the
"unofficial results by precinct" Excel export.

Source: Erie County 825361_5-26-2026_unofficial_results_by_precinct.xlsx

The Excel file has one sheet per contest (plus a turnout summary in Sheet1).
Each contest sheet has a wide layout: row 2 carries the contest header
(e.g. ``GOVERNOR-DEM (Vote for 1) DEM``), row 4 names candidate columns
(``JOSH SHAPIRO\\n(DEM)``, ``robert gannon\\nQualified Write In``, etc.), and
rows 7+ hold per-precinct blocks with four method rows each
(``Election Day``, ``Mail-In``, ``Provisional``, ``Total``).

Per-precinct committee races ("ERIE WARD <ward> DISTRICT <dist> DEMOCRATIC
COMMITTEEMAN" etc.) are skipped. Qualified write-in candidates are aggregated
into a single "Write-In Totals" row per (precinct, office, district, party).

Usage:
    uv run python parsers/pa_erie_primary_2026_precinct_parser.py <input.xlsx> <output.csv>
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

import openpyxl

VOTE_FOR_RE = re.compile(r"\(Vote for\s+(\d+)\)", re.IGNORECASE)
PARTY_SUFFIX_RE = re.compile(r"-(DEM|REP|GP|LBR|IND|GRN|WEP|WFP|PGH|CON)\s*$",
                              re.IGNORECASE)
DISTRICT_ORDINAL_RE = re.compile(
    r"\b(\d+)(?:ST|ND|RD|TH)\s+(?:LEGISLATIVE\s+|CONGRESSIONAL\s+|SENATORIAL\s+)?"
    r"DIS(?:TRICT|T)?\b",
    re.IGNORECASE,
)

STATEWIDE_OFFICES: dict[str, tuple[str, bool]] = {
    "GOVERNOR": ("Governor", False),
    "LIEUTENANT GOVERNOR": ("Lieutenant Governor", False),
    "ATTORNEY GENERAL": ("Attorney General", False),
    "AUDITOR GENERAL": ("Auditor General", False),
    "STATE TREASURER": ("State Treasurer", False),
    "REPRESENTATIVE IN CONGRESS": ("U.S. House", True),
    "SENATOR IN THE GENERAL ASSEMBLY": ("State Senate", True),
    "REPRESENTATIVE IN THE GENERAL ASSEMBLY": ("State House", True),
    "MEMBER OF DEMOCRATIC STATE COMMITTEE": ("Member of Democratic State Committee", False),
    "MEMBER OF REPUBLICAN STATE COMMITTEE": ("Member of Republican State Committee", False),
    "MEMBER OF THE DEMOCRATIC STATE COMMITTEE": ("Member of Democratic State Committee", False),
    "MEMBER OF THE REPUBLICAN STATE COMMITTEE": ("Member of Republican State Committee", False),
    "DEMOCRATIC STATE COMMITTEE": ("Member of Democratic State Committee", False),
    "REPUBLICAN STATE COMMITTEE": ("Member of Republican State Committee", False),
}

# Per-precinct committee race: contest header contains "COMMITTEEMAN" /
# "COMMITTEEWOMAN" / "COMMITTEE PERSON". Does NOT match "STATE COMMITTEE"
# (the countywide Member of State Committee contest).
PER_PRECINCT_COMMITTEE_RE = re.compile(
    r"COMMITTEE\s*(?:MAN|WOMAN|PERSON)S?\b"
    r"|COUNTY\s+COMMITTEE\b",
    re.IGNORECASE,
)

FIELDNAMES = [
    "county", "precinct", "office", "district", "party", "candidate",
    "votes", "election_day", "provisional", "absentee",
]

_ROMAN_RE = re.compile(r"^[IVX]+$")


def _finalize_candidate(raw: str) -> str:
    s = raw.replace(",", "").strip()
    out = []
    for w in s.split():
        if _ROMAN_RE.match(w.upper()):
            out.append(w.upper())
        elif w.upper() in ("JR", "SR"):
            out.append(w.upper().replace("JR", "Jr.").replace("SR", "Sr."))
        elif len(w) >= 3 and w[:2].lower() == "mc":
            out.append("Mc" + w[2:].capitalize())
        else:
            out.append(w.capitalize())
    return " ".join(out)


def _normalize_office(raw: str) -> tuple[str, str]:
    upper = re.sub(r"\s+", " ", raw.upper()).strip()
    dm = DISTRICT_ORDINAL_RE.search(upper)
    district = str(int(dm.group(1))) if dm else ""
    key = DISTRICT_ORDINAL_RE.sub("", upper).strip() if dm else upper
    key = re.sub(r"\s+", " ", key).strip()
    if key in STATEWIDE_OFFICES:
        norm, extract = STATEWIDE_OFFICES[key]
        return (norm, district if extract else "")
    for k, (norm, extract) in STATEWIDE_OFFICES.items():
        if key == k or key.startswith(k + " "):
            return (norm, district if extract else "")
    return (raw.title(), district)


def _parse_contest_header(header: str) -> tuple[str, str, str] | None:
    """Return (office, district, party) from a sheet's contest header row.

    Header looks like ``GOVERNOR-DEM (Vote for  1) \\nDEM  ``. Party is the
    ``-DEM``/``-REP`` suffix on the office; for Member of State Committee
    contests (no suffix), party is derived from the office name.
    """
    if not VOTE_FOR_RE.search(header):
        return None
    # Strip "(Vote for N)" and any trailing party line.
    body = VOTE_FOR_RE.sub("", header).strip()
    body = body.split("\n")[0].strip()
    psm = PARTY_SUFFIX_RE.search(body)
    if psm:
        party = psm.group(1).upper()
        office_text = PARTY_SUFFIX_RE.sub("", body).strip()
    else:
        office_text = body
        if "DEMOCRATIC" in office_text.upper():
            party = "DEM"
        elif "REPUBLICAN" in office_text.upper():
            party = "REP"
        else:
            party = ""
    if PER_PRECINCT_COMMITTEE_RE.search(office_text):
        return None
    office, district = _normalize_office(office_text)
    return (office, district, party)


def _is_candidate_header(cell_value) -> bool:
    """Return True if a row-4 cell names a candidate column.

    Candidate columns look like ``JOSH SHAPIRO\\n(DEM) `` or
    ``robert gannon\\nQualified Write In \\n`` or ``Unresolved\\nWrite-In``.
    """
    if not cell_value or not isinstance(cell_value, str):
        return False
    s = cell_value.strip()
    if not s:
        return False
    # Skip the column-label cells that aren't candidates.
    if s in ("Precinct", "Times Cast", "Registered\nVoters",
             "Registered Voters", "Undervotes", "County", "Total Votes",
             "Overvotes"):
        return False
    # Candidate cells contain either "(DEM)"/"(REP)" party tag or
    # "Qualified Write In" or "Write-In".
    if "(DEM)" in s or "(REP)" in s or "Qualified Write In" in s or "Write-In" in s:
        return True
    return False


def _candidate_name_from_header(cell_value: str) -> tuple[str, bool]:
    """Return (display_name, is_writein) for a candidate header cell."""
    s = cell_value.strip()
    # "JOSH SHAPIRO\n(DEM)" -> "JOSH SHAPIRO"
    # "robert gannon\nQualified Write In" -> "robert gannon" (write-in)
    # "Unresolved\nWrite-In" -> write-in aggregate
    first_line = s.split("\n")[0].strip()
    if first_line.upper() in ("UNRESOLVED", "WRITE-IN", "WRITE IN"):
        return ("Write-In Totals", True)
    is_writein = "Qualified Write In" in s or "WRITE-IN" in s.upper()
    # Strip "(DEM)"/"(REP)" suffix if present.
    name = re.sub(r"\s*\((?:DEM|REP|GP|LBR|IND|GRN|WEP|WFP|PGH|CON)\)\s*$",
                  "", first_line).strip()
    return (name, is_writein)


def parse_sheet(ws, county: str) -> list[dict]:
    """Parse one contest sheet into per-precinct candidate rows."""
    # Contest header is in row 2, column 1.
    header = ws.cell(2, 1).value
    if not header or not isinstance(header, str):
        return []
    parsed = _parse_contest_header(header)
    if parsed is None:
        return []
    office, district, party = parsed

    # Row 4 holds column headers. Identify candidate column indices.
    # Each candidate has a single data column (the column under their name).
    candidate_cols: list[tuple[int, str, bool]] = []  # (col_idx, name, is_writein)
    row4_max = ws.max_column
    for c in range(1, row4_max + 1):
        val = ws.cell(4, c).value
        if _is_candidate_header(val):
            name, is_writein = _candidate_name_from_header(val)
            candidate_cols.append((c, name, is_writein))
    if not candidate_cols:
        return []

    rows: list[dict] = []
    current_precinct: str | None = None
    # Per-precinct method accumulators per candidate column.
    # method_values[col_idx] = {"ed": int, "mi": int, "pr": int, "total": int}
    method_values: dict[int, dict] | None = None
    precinct_methods: dict[int, dict] = {c: {"ed": 0, "mi": 0, "pr": 0}
                                          for c, _, _ in candidate_cols}

    # Iterate rows starting at 7. Each precinct has a header row
    # (precinct name in col 1) followed by 4 method rows.
    r = 7
    n = ws.max_row + 1
    while r < n:
        col1 = ws.cell(r, 1).value
        if col1 and isinstance(col1, str):
            label = col1.strip()
            if label in ("Election Day", "Mail-In", "Provisional", "Total"):
                # Method row for the current precinct.
                if current_precinct is None:
                    r += 1
                    continue
                if method_values is None:
                    method_values = {c: {"ed": 0, "mi": 0, "pr": 0}
                                     for c, _, _ in candidate_cols}
                if label == "Election Day":
                    for cidx, _, _ in candidate_cols:
                        v = ws.cell(r, cidx).value or 0
                        method_values[cidx]["ed"] += int(v) if isinstance(v, (int, float)) else 0
                elif label == "Mail-In":
                    for cidx, _, _ in candidate_cols:
                        v = ws.cell(r, cidx).value or 0
                        method_values[cidx]["mi"] += int(v) if isinstance(v, (int, float)) else 0
                elif label == "Provisional":
                    for cidx, _, _ in candidate_cols:
                        v = ws.cell(r, cidx).value or 0
                        method_values[cidx]["pr"] += int(v) if isinstance(v, (int, float)) else 0
                elif label == "Total":
                    # Emit one row per candidate for this precinct.
                    writein_agg: dict[tuple, dict] = {}
                    for cidx, name, is_writein in candidate_cols:
                        total_val = ws.cell(r, cidx).value or 0
                        total = int(total_val) if isinstance(total_val, (int, float)) else 0
                        ed = method_values[cidx]["ed"] if method_values else 0
                        mi = method_values[cidx]["mi"] if method_values else 0
                        pr = method_values[cidx]["pr"] if method_values else 0
                        if is_writein:
                            key = (current_precinct, office, district, party)
                            agg = writein_agg.setdefault(key, {
                                "county": county,
                                "precinct": current_precinct,
                                "office": office,
                                "district": district,
                                "party": party,
                                "candidate": "Write-In Totals",
                                "votes": 0, "election_day": 0,
                                "provisional": 0, "absentee": 0,
                            })
                            agg["votes"] += total
                            agg["election_day"] += ed
                            agg["absentee"] += mi
                            agg["provisional"] += pr
                        else:
                            candidate = _finalize_candidate(name)
                            rows.append({
                                "county": county,
                                "precinct": current_precinct,
                                "office": office,
                                "district": district,
                                "party": party,
                                "candidate": candidate,
                                "votes": total,
                                "election_day": ed,
                                "provisional": pr,
                                "absentee": mi,
                            })
                    rows.extend(writein_agg.values())
                    method_values = None
                r += 1
                continue
            # Otherwise it's a precinct name (skip "County"/"Erie County").
            if label in ("County", "Erie County"):
                r += 1
                continue
            current_precinct = label
            method_values = {c: {"ed": 0, "mi": 0, "pr": 0}
                             for c, _, _ in candidate_cols}
            r += 1
            continue
        r += 1

    return rows


def parse_workbook(xlsx_path: Path, county: str = "Erie") -> list[dict]:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    all_rows: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Sheet1 is the county-wide turnout summary; skip.
        if ws.cell(2, 1).value is None:
            continue
        sheet_rows = parse_sheet(ws, county)
        all_rows.extend(sheet_rows)
    return all_rows


def main(argv: list[str]) -> None:
    if len(argv) != 3:
        sys.exit(f"Usage: {Path(argv[0]).name} <input.xlsx> <output.csv>")
    xlsx_path = Path(argv[1])
    out_path = Path(argv[2])
    if not xlsx_path.exists():
        sys.exit(f"Missing XLSX: {xlsx_path}")
    rows = parse_workbook(xlsx_path)
    with out_path.open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDNAMES)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    offices = len({(r["office"], r["district"]) for r in rows})
    precincts = len({r["precinct"] for r in rows})
    print(f"Wrote {len(rows)} rows across {offices} contests / "
          f"{precincts} precincts to {out_path}")


if __name__ == "__main__":
    main(sys.argv)